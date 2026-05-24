"""Round-8 sweep summary -- PER AREA-PAIR, all IFI lag-windows, full stats.

For every config in a sweep, reports the headline metrics for each of the 10
area-pairs (learner cohort). The statistical unit is the significant
communication-subspace dimension, pooled across a pair's animals (never across
pairs); ``n_dims`` makes that n explicit.

Writes, into figures/:
  * sweep_summary_<name>.csv   long format -- one row per (config x pair),
                               every metric incl. IFI at lag-windows 1..10
  * sweep_summary_<name>.xlsx  sheet 'robustness' (per-pair, how often each
                               effect holds across the sweep -- colour-scaled)
                               + sheet 'data' (the long table, p-columns red)
  * sweep_pairs_<name>_*.png   per-pair grids across configs
  * sweep_ifi_windows_<name>.png   per-pair IFI vs lag-window

Run:  python scripts/summarise_sweep.py --sweep spatial
"""

from __future__ import annotations

import argparse
import pickle
import sys
from pathlib import Path

import numpy as np

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

import matplotlib  # noqa: E402

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import openpyxl  # noqa: E402
from openpyxl.formatting.rule import ColorScaleRule  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from scipy import stats  # noqa: E402

from striatum_cca import config, sweep  # noqa: E402

ALPHA = 0.05
EPOCHS = ("naive", "expert")
N_WINDOWS = 10
PAIR_NAMES = [f"{ax}-{ay}" for ax, ay in config.PAIRS]
IFI_KEYS = [f"ifi_w{w}" for w in range(1, N_WINDOWS + 1)]
P_IFI_KEYS = [f"p_ifi_w{w}" for w in range(1, N_WINDOWS + 1)]


def _load(tag, stage):
    path = config.RESULTS_DIR / f"stage{stage}_{tag}.pkl"
    if not path.exists():
        return []
    with open(path, "rb") as fh:
        return pickle.load(fh)["results"]


def _sig(ea):
    return np.where(ea.p_per_dim < ALPHA)[0]


def _clean(values):
    a = np.asarray(values, float)
    return a[np.isfinite(a)]


def _wilcoxon(v):
    return (float(stats.wilcoxon(v).pvalue)
            if v.size >= 6 and np.any(v != 0) else np.nan)


def _krule_label(cfg) -> str:
    if cfg.k_mode == "fixed":
        return f"fixed{cfg.k_fixed}"
    if cfg.k_mode == "variance":
        return f"var{int(round(cfg.k_variance * 100))}"
    return f"samples{cfg.samples_per_pc}"


def cfg_params(cfg) -> dict:
    if cfg.bin_mode == "temporal":
        binning = f"{cfg.temporal_bin_ms}ms"
    else:
        binning = "5cm" if cfg.max_lag_bins == 5 else "2.5cm"
    return {
        "bin": binning,
        "cca": "residual" if cfg.subtract_trial_mean else "signal",
        "fs": "excl" if cfg.exclude_fast_spiking else "incl",
        "z": "on" if cfg.zscore_units else "off",
        "k_rule": _krule_label(cfg),
        "min_units": cfg.min_units,
        "lp_consec": cfg.lp_min_consecutive,
    }


def pair_metrics(s2p, s3p) -> dict:
    """Metrics for ONE area-pair (its learner animals only). The statistical
    unit is the significant subspace dimension, pooled over the pair's animals.
    """
    cc = {e: [] for e in EPOCHS}
    ifi_by_w = {w: [] for w in range(1, N_WINDOWS + 1)}
    for r in s2p:
        for e in EPOCHS:
            ea = r.epochs[e]
            js = _sig(ea)
            cc[e].extend(ea.held_out_cc[js].tolist())
            n_w = ea.ifi_windows.shape[1]
            for w in range(1, N_WINDOWS + 1):
                if w - 1 < n_w:
                    ifi_by_w[w].extend(ea.ifi_windows[js, w - 1].tolist())
    cc_n, cc_e = _clean(cc["naive"]), _clean(cc["expert"])
    m: dict = {"n_learn": len(s2p),
               "n_dims_naive": cc_n.size, "n_dims_expert": cc_e.size}
    m["cc_naive"] = float(np.mean(cc_n)) if cc_n.size else np.nan
    m["cc_expert"] = float(np.mean(cc_e)) if cc_e.size else np.nan
    m["d_cc"] = m["cc_expert"] - m["cc_naive"]
    m["p_naive_vs_expert"] = (float(stats.mannwhitneyu(cc_n, cc_e).pvalue)
                              if cc_n.size >= 3 and cc_e.size >= 3 else np.nan)
    for w in range(1, N_WINDOWS + 1):
        v = _clean(ifi_by_w[w])
        m[f"ifi_w{w}"] = float(np.mean(v)) if v.size else np.nan
        m[f"p_ifi_w{w}"] = _wilcoxon(v)
    ang = [r.angles_x["naive->expert"].max() for r in s3p]
    floor = [np.nanmean([r.epochs[e].split_half_angle_x.max() for e in EPOCHS])
             for r in s3p]
    m["angle_ne"] = float(np.nanmean(ang)) if ang else np.nan
    m["sh_floor"] = float(np.nanmean(floor)) if floor else np.nan
    m["angle_minus_floor"] = m["angle_ne"] - m["sh_floor"]
    m["gini_naive"] = (float(np.nanmean([r.epochs["naive"].gini_x for r in s3p]))
                       if s3p else np.nan)
    m["gini_expert"] = (
        float(np.nanmean([r.epochs["expert"].gini_x for r in s3p]))
        if s3p else np.nan)
    return m


def metrics_for(tag) -> dict:
    s2, s3 = _load(tag, 2), _load(tag, 3)
    out = {}
    for (ax, ay), name in zip(config.PAIRS, PAIR_NAMES):
        s2p = [r for r in s2
               if (r.area_x, r.area_y) == (ax, ay) and r.role == "learner"]
        s3p = [r for r in s3
               if (r.area_x, r.area_y) == (ax, ay) and r.role == "learner"]
        out[name] = pair_metrics(s2p, s3p)
    return out


METRIC_COLS = (["n_learn", "n_dims_naive", "n_dims_expert", "cc_naive",
                "cc_expert", "d_cc", "p_naive_vs_expert", "angle_ne",
                "sh_floor", "angle_minus_floor", "gini_naive", "gini_expert"]
               + IFI_KEYS + P_IFI_KEYS)
PARAM_COLS = ["tag", "bin", "cca", "fs", "z", "k_rule", "min_units",
              "lp_consec", "pair"]
LONG_COLS = PARAM_COLS + METRIC_COLS
P_COLS = ["p_naive_vs_expert"] + P_IFI_KEYS


def _round(v):
    if isinstance(v, float):
        return None if not np.isfinite(v) else round(v, 4)
    return v


def write_csv(tags, params, data, path):
    with open(path, "w") as fh:
        fh.write(",".join(LONG_COLS) + "\n")
        for tag in tags:
            for name in PAIR_NAMES:
                row = {"tag": tag, "pair": name, **params[tag],
                       **data[tag][name]}
                fh.write(",".join("" if _round(row.get(c)) is None
                                  else str(_round(row.get(c, "")))
                                  for c in LONG_COLS) + "\n")
    print(f"saved {path}")


def _red_scale(ws, rng):
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=0, start_color="C6584B",
        mid_type="num", mid_value=0.05, mid_color="F5B7B1",
        end_type="num", end_value=0.2, end_color="FFFFFF"))


def _diverge(ws, rng):
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="F8696B", mid_type="percentile",
        mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"))


def write_xlsx(tags, params, data, path, sweep_name):
    wb = openpyxl.Workbook()

    # --- sheet 1: per-pair robustness across the whole sweep ----------------
    ws = wb.active
    ws.title = "robustness"
    ws["A1"] = (f"Round-8 {sweep_name} sweep -- per-pair robustness across "
                f"{len(tags)} configs. n = significant subspace dimensions "
                f"(pooled within each pair). Counts are configs (out of "
                f"{len(tags)}) in which the effect is present / significant.")
    ws["A1"].font = Font(bold=True, size=11)
    ws.merge_cells("A1:R1")
    hdr = (["pair", "median n_dims", "median dCC", "dCC>0 configs",
            "strength p<.05 configs", "median angle-floor",
            "angle>floor configs"]
           + [f"IFI w{w} p<.05" for w in range(1, N_WINDOWS + 1)])
    ws.append(hdr)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for name in PAIR_NAMES:
        rows = [data[t][name] for t in tags]
        nd = _clean([0.5 * (r["n_dims_naive"] + r["n_dims_expert"])
                     for r in rows])
        dcc = _clean([r["d_cc"] for r in rows])
        pne = _clean([r["p_naive_vs_expert"] for r in rows])
        amf = _clean([r["angle_minus_floor"] for r in rows])
        line = [name,
                float(np.median(nd)) if nd.size else None,
                float(np.median(dcc)) if dcc.size else None,
                int(np.sum(dcc > 0)),
                int(np.sum(pne < ALPHA)),
                float(np.median(amf)) if amf.size else None,
                int(np.sum(amf > 0))]
        for w in range(1, N_WINDOWS + 1):
            pw = _clean([r[f"p_ifi_w{w}"] for r in rows])
            line.append(int(np.sum(pw < ALPHA)))
        ws.append([_round(v) for v in line])
    last = len(PAIR_NAMES) + 2
    _diverge(ws, f"D3:D{last}")                       # dCC>0 count
    _diverge(ws, f"G3:G{last}")                       # angle>floor count
    for col in range(8, 8 + N_WINDOWS):               # IFI sig-count columns
        c = get_column_letter(col)
        _diverge(ws, f"{c}3:{c}{last}")
    ws.column_dimensions["A"].width = 11
    for col in range(2, 8 + N_WINDOWS):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.freeze_panes = "B3"

    # --- sheet 2: the full long-format table -------------------------------
    ds = wb.create_sheet("data")
    ds.append(LONG_COLS)
    for cell in ds[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=8)
        cell.fill = PatternFill("solid", fgColor="404040")
    for tag in tags:
        for name in PAIR_NAMES:
            row = {"tag": tag, "pair": name, **params[tag], **data[tag][name]}
            ds.append([_round(row.get(c)) for c in LONG_COLS])
    n_rows = len(tags) * len(PAIR_NAMES) + 1
    for pcol in P_COLS:                               # colour every p column
        c = get_column_letter(LONG_COLS.index(pcol) + 1)
        _red_scale(ds, f"{c}2:{c}{n_rows}")
    ds.freeze_panes = "A2"
    wb.save(path)
    print(f"saved {path}  ({len(tags)} configs x {len(PAIR_NAMES)} pairs)")


def pair_grid(tags, data, key, label, ref, title, path):
    fig, axes = plt.subplots(2, 5, figsize=(16, 6.8))
    axes = axes.ravel()
    x = np.arange(len(tags))
    for ax, name in zip(axes, PAIR_NAMES):
        vals = np.array([data[t][name].get(key, np.nan) for t in tags], float)
        if ref is not None:
            ax.axhline(ref, color="k", ls="--", lw=0.8)
        ax.scatter(x, vals, s=10, color="tab:blue", alpha=0.5)
        ax.set_title(f"{name}  ({int(np.sum(np.isfinite(vals)))}/{len(tags)})",
                     fontsize=9)
        ax.set_xticks([])
    for ax in axes[::5]:
        ax.set_ylabel(label, fontsize=8)
    for ax in axes[5:]:
        ax.set_xlabel(f"config (of {len(tags)})", fontsize=8)
    fig.suptitle(title, fontsize=11)
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f"saved {path}")


def ifi_window_figure(tags, data, path):
    fig, axes = plt.subplots(2, 5, figsize=(16, 6.8))
    axes = axes.ravel()
    w = np.arange(1, N_WINDOWS + 1)
    for ax, name in zip(axes, PAIR_NAMES):
        mat = np.array([[data[t][name].get(f"ifi_w{i}", np.nan)
                         for i in w] for t in tags], float)
        mean = np.nanmean(mat, axis=0)
        sd = np.nanstd(mat, axis=0)
        ax.axhline(0, color="k", ls="--", lw=0.8)
        ax.errorbar(w, mean, yerr=sd, fmt="o-", color="tab:purple",
                    ms=4, lw=1.3, capsize=2)
        ax.set_title(name, fontsize=9)
        ax.set_ylim(-1.05, 1.05)
    for ax in axes[::5]:
        ax.set_ylabel("IFI (mean +/- SD over configs)", fontsize=8)
    for ax in axes[5:]:
        ax.set_xlabel("lag-integration window (bins)", fontsize=8)
    fig.suptitle("Information Flow Index per lag-window, per pair "
                 "(mean +/- SD across the sweep; dashed = no net flow)",
                 fontsize=11)
    fig.tight_layout()
    fig.savefig(path, dpi=150)
    plt.close(fig)
    print(f"saved {path}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--sweep", choices=("spatial", "temporal"),
                   default="spatial")
    name = p.parse_args().sweep
    grid = sweep.build_sweep(name)
    tags = [t for t, _, _ in grid]
    have = [t for t in tags
            if (config.RESULTS_DIR / f"stage2_{t}.pkl").exists()]
    if not have:
        print("no stage2 results found for this sweep -- run run_stage2.py first")
        return
    if len(have) < len(tags):
        print(f"note: {len(have)}/{len(tags)} configs have results; "
              f"summarising those.")
    tags = have
    params = {t: dict(tag=t, **cfg_params(cfg))
              for t, _, cfg in grid if t in have}
    data = {t: metrics_for(t) for t in tags}

    fd = config.FIGURES_DIR
    fd.mkdir(parents=True, exist_ok=True)
    write_csv(tags, params, data, fd / f"sweep_summary_{name}.csv")
    write_xlsx(tags, params, data, fd / f"sweep_summary_{name}.xlsx", name)
    pair_grid(tags, data, "d_cc", "dCC (expert-naive)", 0.0,
              f"Communication-strength change across learning -- {name} sweep "
              f"(dashed = no change; tight cluster = robust)",
              fd / f"sweep_pairs_{name}_d_cc.png")
    pair_grid(tags, data, "p_naive_vs_expert", "naive-vs-expert p", 0.05,
              f"Strength naive-vs-expert significance -- {name} sweep "
              f"(dashed = 0.05; below = significant)",
              fd / f"sweep_pairs_{name}_p.png")
    pair_grid(tags, data, "angle_minus_floor", "angle - split-half floor (rad)",
              0.0, f"Subspace reorientation above the noise floor -- {name} "
              f"sweep (dashed = at floor; above = reorientation)",
              fd / f"sweep_pairs_{name}_angle.png")
    ifi_window_figure(tags, data, fd / f"sweep_ifi_windows_{name}.png")

    n = len(tags)
    print(f"\nPer-pair robustness across {n} {name} configs:")
    print(f"  {'pair':>9} {'med n_dims':>11} {'dCC>0':>9} {'p<.05':>8} "
          f"{'angle>floor':>12}")
    for name_ in PAIR_NAMES:
        rows = [data[t][name_] for t in tags]
        nd = _clean([0.5 * (r["n_dims_naive"] + r["n_dims_expert"])
                     for r in rows])
        dcc = _clean([r["d_cc"] for r in rows])
        pne = _clean([r["p_naive_vs_expert"] for r in rows])
        amf = _clean([r["angle_minus_floor"] for r in rows])
        print(f"  {name_:>9} {np.median(nd) if nd.size else 0:>11.1f} "
              f"{int(np.sum(dcc > 0)):>5}/{n} {int(np.sum(pne < ALPHA)):>4}/{n} "
              f"{int(np.sum(amf > 0)):>7}/{n}")


if __name__ == "__main__":
    main()
